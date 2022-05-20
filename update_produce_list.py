import tkinter as tk
import openpyxl

# ファイル選択ダイアログの表示
def fileselect_button_clicked():
  print('変更ボタンがクリックされました')

# 更新ボタンがクリックされた時の動作
def update_button_clicked():
  # 更新用データの用意
  for i in range(5):
    produce_name = name_list[i].get().replace('\x10', '')
    price_str = price_list[i].get().replace('\x10', '')

    # 入力が空でない場合のみデータを追加
    if produce_name != ""  and price_str != "":
      price = int(price_str)
      update_dict[produce_name] = price

  # エクセルファイルの更新
  wb = openpyxl.load_workbook('全体売上.xlsx')
  sheet = wb.worksheets[0]

  for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in update_dict:
      sheet.cell(row=row_num, column=2).value = update_dict[produce_name]
      fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FF6E91') 
      sheet.cell(row=row_num, column=2).fill = fill

  wb.save('価格更新.xlsx')

# 初期設定
name_list = []
price_list = []
update_dict = {}

# アプリの雛形作成
root = tk.Tk()
root.title('農産物価格更新')
root.geometry('450x650')

# ファイル名の表示
filename_labelframe = tk.LabelFrame(root, text = " 更新ファイル名 ", font=("Meiryo",16))
filename_labelframe.place(x=50, y=60, width=350, height=80)

filename_label = tk.Label(filename_labelframe, text='全体売上.xlsx', font=("Meiryo",14))
filename_label.place(x=20, y=10)

file_select_button = tk.Button(filename_labelframe, text='変更', font=("Meiryo",16), command=fileselect_button_clicked)
file_select_button.place(x=270, y=10, width=60, height=30)

# ラベルの作成
name_label = tk.Label(root, text='農産物の名称', font=("Meiryo",16))
name_label.place(x=50 , y=200)

price_label = tk.Label(root, text='変更後の価格', font=("Meiryo",16))
price_label.place(x=260, y=200)

# テキストボックスの作成
for i in range(5):
  text_box = tk.Entry(root, font=("Meiryo",14))
  text_box.place(x=50, y=(235 + i * 55), width=185, height=40)
  name_list.append(text_box)

  text_box2 = tk.Entry(root, font=("Meiryo",14))
  text_box2.place(x=260, y=(235 + i * 55), width=140, height=40)
  price_list.append(text_box2)

# 更新ボタンの作成
update_button = tk.Button(root, text='更新', font=("Meiryo",18), command=update_button_clicked)
update_button.place(x=135, y=565, width=180, height=45)

# アプリの実行
root.mainloop()


